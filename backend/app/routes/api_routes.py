"""Main API routes for CRUD operations on all entities."""

import os
import uuid
from datetime import date, datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, File, UploadFile, Form, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel

from app.database import get_db, SessionLocal
from app.models import (
    User, UserRole, Fleet, Bus, Cylinder, CylinderStatus,
    Inspection, InspectionType, InspectionResult,
    Repair, StockItem, StockTransaction, TransactionType,
    NotificationSettings,
)
from app.auth import get_current_user, require_role, get_user_fleet_id

BASE_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

router = APIRouter()


# ── Helpers ──────────────────────────────────────────────────────────────────

def scope_query(db, model, user):
    """Scope query to user's fleet if user is not admin."""
    q = db.query(model)
    if user.role != UserRole.ADMIN and user.fleet_id:
        q = q.filter(model.fleet_id == user.fleet_id)
    return q


def save_upload(file: UploadFile, prefix: str = "") -> str:
    """Save uploaded file and return relative path."""
    uploads_dir = os.path.join(BASE_DIR, "uploads")
    os.makedirs(uploads_dir, exist_ok=True)
    ext = os.path.splitext(file.filename or "")[1] or ".jpg"
    filename = f"{prefix}_{uuid.uuid4().hex}{ext}"
    filepath = os.path.join(uploads_dir, filename)
    with open(filepath, "wb") as f:
        f.write(file.file.read())
    return f"/uploads/{filename}"


# ── Fleets ───────────────────────────────────────────────────────────────────

class FleetCreate(BaseModel):
    name: str
    address: str | None = None
    responsible: str | None = None
    comment: str | None = None


@router.get("/fleets")
def list_fleets(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """List all fleets. Admins see all; others see only their fleet."""
    if user.role == UserRole.ADMIN:
        fleets = db.query(Fleet).all()
    else:
        fleets = db.query(Fleet).filter(Fleet.id == user.fleet_id).all()
    return fleets


@router.post("/fleets")
def create_fleet(
    data: FleetCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN)),
):
    fleet = Fleet(**data.model_dump())
    db.add(fleet)
    db.commit()
    db.refresh(fleet)
    return fleet


@router.put("/fleets/{fleet_id}")
def update_fleet(
    fleet_id: int,
    data: FleetCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN)),
):
    fleet = db.query(Fleet).filter(Fleet.id == fleet_id).first()
    if not fleet:
        raise HTTPException(404, "Автопарк не найден")
    for k, v in data.model_dump().items():
        setattr(fleet, k, v)
    db.commit()
    return fleet


@router.delete("/fleets/{fleet_id}")
def delete_fleet(
    fleet_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN)),
):
    fleet = db.query(Fleet).filter(Fleet.id == fleet_id).first()
    if not fleet:
        raise HTTPException(404)
    db.delete(fleet)
    db.commit()
    return {"ok": True}


# ── Buses ────────────────────────────────────────────────────────────────────

class BusCreate(BaseModel):
    fleet_id: int
    gosnomer: str
    board_number: str | None = None
    vin: str | None = None
    model: str | None = None
    year: int | None = None


@router.get("/buses")
def list_buses(
    fleet_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = scope_query(db, Bus, user)
    if fleet_id:
        q = q.filter(Bus.fleet_id == fleet_id)
    return q.all()


@router.post("/buses")
def create_bus(
    data: BusCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    bus = Bus(**data.model_dump())
    db.add(bus)
    db.commit()
    db.refresh(bus)
    return bus


@router.put("/buses/{bus_id}")
def update_bus(bus_id: int, data: BusCreate, db: Session = Depends(get_db),
               user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER))):
    bus = db.query(Bus).filter(Bus.id == bus_id).first()
    if not bus:
        raise HTTPException(404)
    for k, v in data.model_dump().items():
        setattr(bus, k, v)
    db.commit()
    return bus


@router.delete("/buses/{bus_id}")
def delete_bus(bus_id: int, db: Session = Depends(get_db),
               user: User = Depends(require_role(UserRole.ADMIN))):
    bus = db.query(Bus).filter(Bus.id == bus_id).first()
    if not bus:
        raise HTTPException(404)
    db.delete(bus)
    db.commit()
    return {"ok": True}


# ── Cylinders ────────────────────────────────────────────────────────────────

class CylinderCreate(BaseModel):
    fleet_id: int
    bus_id: int | None = None
    number: str
    stamp: str | None = None
    serial_number: str | None = None
    gas_type: str | None = None
    manufactured_date: str | None = None
    capacity_liters: float | None = None
    working_pressure: float | None = None
    test_pressure: float | None = None
    tare_weight: float | None = None
    status: CylinderStatus = CylinderStatus.ACTIVE
    next_inspection_date: str | None = None


@router.get("/cylinders")
def list_cylinders(
    fleet_id: int | None = None,
    bus_id: int | None = None,
    status: CylinderStatus | None = None,
    search: str | None = None,
    overdue: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = scope_query(db, Cylinder, user)
    if fleet_id:
        q = q.filter(Cylinder.fleet_id == fleet_id)
    if bus_id:
        q = q.filter(Cylinder.bus_id == bus_id)
    if status:
        q = q.filter(Cylinder.status == status)
    if search:
        q = q.filter(
            (Cylinder.number.contains(search)) |
            (Cylinder.stamp.contains(search)) |
            (Cylinder.serial_number.contains(search))
        )
    if overdue:
        q = q.filter(
            Cylinder.next_inspection_date < date.today(),
            Cylinder.status == CylinderStatus.ACTIVE,
        )
    return q.all()


@router.post("/cylinders")
def create_cylinder(
    data: CylinderCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    d = data.model_dump()
    if d.get("manufactured_date"):
        d["manufactured_date"] = date.fromisoformat(d["manufactured_date"])
    if d.get("next_inspection_date"):
        d["next_inspection_date"] = date.fromisoformat(d["next_inspection_date"])
    cyl = Cylinder(**d)
    db.add(cyl)
    db.commit()
    db.refresh(cyl)
    return cyl


@router.put("/cylinders/{cyl_id}")
def update_cylinder(
    cyl_id: int,
    data: CylinderCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    cyl = db.query(Cylinder).filter(Cylinder.id == cyl_id).first()
    if not cyl:
        raise HTTPException(404)
    d = data.model_dump()
    if d.get("manufactured_date"):
        d["manufactured_date"] = date.fromisoformat(d["manufactured_date"])
    if d.get("next_inspection_date"):
        d["next_inspection_date"] = date.fromisoformat(d["next_inspection_date"])
    for k, v in d.items():
        setattr(cyl, k, v)
    db.commit()
    return cyl


@router.post("/cylinders/{cyl_id}/photo")
def upload_cylinder_photo(
    cyl_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    cyl = db.query(Cylinder).filter(Cylinder.id == cyl_id).first()
    if not cyl:
        raise HTTPException(404)
    path = save_upload(file, f"cyl_{cyl_id}")
    cyl.photo_path = path
    db.commit()
    return {"photo_path": path}


# ── Inspections ──────────────────────────────────────────────────────────────

class InspectionCreate(BaseModel):
    cylinder_id: int
    inspection_date: str
    inspection_type: InspectionType = InspectionType.STANDARD
    result: InspectionResult = InspectionResult.PENDING
    next_inspection_date: str | None = None
    inspector: str | None = None
    visual_inspection: bool = False
    hydraulic_test: bool = False
    pneumatic_test: bool = False
    weight_check: bool = False
    ultrasonic_thickness: bool = False
    defectoscopy: bool = False
    powder_test: bool = False
    magnetic_test: bool = False
    pressure_achieved: float | None = None
    wall_thickness: float | None = None
    weight_measured: float | None = None
    notes: str | None = None


@router.get("/inspections")
def list_inspections(
    cylinder_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = db.query(Inspection)
    if cylinder_id:
        q = q.filter(Inspection.cylinder_id == cylinder_id)
    # Scope to user's fleet
    if user.role != UserRole.ADMIN and user.fleet_id:
        q = q.join(Cylinder).filter(Cylinder.fleet_id == user.fleet_id)
    return q.order_by(Inspection.inspection_date.desc()).all()


@router.post("/inspections")
def create_inspection(
    data: InspectionCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    d = data.model_dump()
    d["inspection_date"] = date.fromisoformat(d["inspection_date"])
    if d.get("next_inspection_date"):
        d["next_inspection_date"] = date.fromisoformat(d["next_inspection_date"])
    insp = Inspection(**d)
    db.add(insp)

    # Update cylinder's next_inspection_date
    if insp.next_inspection_date:
        cyl = db.query(Cylinder).filter(Cylinder.id == insp.cylinder_id).first()
        if cyl:
            cyl.next_inspection_date = insp.next_inspection_date
            cyl.status = CylinderStatus.INSPECTION

    db.commit()
    db.refresh(insp)
    return insp


@router.post("/inspections/{insp_id}/photo")
def upload_inspection_photo(
    insp_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    insp = db.query(Inspection).filter(Inspection.id == insp_id).first()
    if not insp:
        raise HTTPException(404)
    path = save_upload(file, f"insp_{insp_id}")
    insp.photo_path = path
    db.commit()
    return {"photo_path": path}


# ── Repairs ──────────────────────────────────────────────────────────────────

class RepairCreate(BaseModel):
    cylinder_id: int | None = None
    bus_id: int | None = None
    fleet_id: int
    repair_date: str = str(date.today())
    description: str
    parts_replaced: str | None = None
    cost: float | None = None
    technician: str | None = None
    notes: str | None = None


@router.get("/repairs")
def list_repairs(
    fleet_id: int | None = None,
    bus_id: int | None = None,
    cylinder_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = scope_query(db, Repair, user)
    if fleet_id:
        q = q.filter(Repair.fleet_id == fleet_id)
    if bus_id:
        q = q.filter(Repair.bus_id == bus_id)
    if cylinder_id:
        q = q.filter(Repair.cylinder_id == cylinder_id)
    return q.order_by(Repair.repair_date.desc()).all()


@router.post("/repairs")
def create_repair(
    data: RepairCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    d = data.model_dump()
    if d.get("repair_date"):
        d["repair_date"] = date.fromisoformat(d["repair_date"])
    repair = Repair(**d)
    db.add(repair)
    db.commit()
    db.refresh(repair)
    return repair


# ── Stock ────────────────────────────────────────────────────────────────────

class StockItemCreate(BaseModel):
    fleet_id: int
    name: str
    category: str | None = None
    quantity: float = 0
    unit: str = "шт"
    min_quantity: float = 0
    price: float | None = None


class StockTransactionCreate(BaseModel):
    stock_item_id: int
    fleet_id: int
    transaction_type: TransactionType
    quantity: float
    comment: str | None = None
    related_repair_id: int | None = None


@router.get("/stock")
def list_stock(
    fleet_id: int | None = None,
    low_stock: bool = False,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = scope_query(db, StockItem, user)
    if fleet_id:
        q = q.filter(StockItem.fleet_id == fleet_id)
    if low_stock:
        q = q.filter(StockItem.quantity <= StockItem.min_quantity)
    return q.all()


@router.post("/stock")
def create_stock_item(
    data: StockItemCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    item = StockItem(**data.model_dump())
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


@router.put("/stock/{item_id}")
def update_stock_item(
    item_id: int,
    data: StockItemCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    item = db.query(StockItem).filter(StockItem.id == item_id).first()
    if not item:
        raise HTTPException(404)
    for k, v in data.model_dump().items():
        setattr(item, k, v)
    db.commit()
    return item


@router.post("/stock/transaction")
def create_stock_transaction(
    data: StockTransactionCreate,
    db: Session = Depends(get_db),
    user: User = Depends(require_role(UserRole.ADMIN, UserRole.WORKER)),
):
    item = db.query(StockItem).filter(StockItem.id == data.stock_item_id).first()
    if not item:
        raise HTTPException(404, "Товар не найден")

    if data.transaction_type == TransactionType.IN:
        item.quantity += data.quantity
    elif data.transaction_type in (TransactionType.OUT, TransactionType.RESERVED):
        if item.quantity < data.quantity:
            raise HTTPException(400, "Недостаточно на складе")
        item.quantity -= data.quantity

    txn = StockTransaction(**data.model_dump(), date=date.today())
    db.add(txn)
    db.commit()
    db.refresh(txn)
    return txn


@router.get("/stock/transactions")
def list_stock_transactions(
    stock_item_id: int | None = None,
    fleet_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = scope_query(db, StockTransaction, user)
    if stock_item_id:
        q = q.filter(StockTransaction.stock_item_id == stock_item_id)
    if fleet_id:
        q = q.filter(StockTransaction.fleet_id == fleet_id)
    return q.order_by(StockTransaction.date.desc()).limit(200).all()


# ── Reports ──────────────────────────────────────────────────────────────────

@router.get("/reports/summary")
def report_summary(
    fleet_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Aggregated report per fleet."""
    fleets_q = scope_query(db, Fleet, user)
    if fleet_id:
        fleets_q = fleets_q.filter(Fleet.id == fleet_id)

    results = []
    for fleet in fleets_q.all():
        buses = db.query(Bus).filter(Bus.fleet_id == fleet.id).count()
        cylinders = db.query(Cylinder).filter(Cylinder.fleet_id == fleet.id).count()
        overdue = db.query(Cylinder).filter(
            Cylinder.fleet_id == fleet.id,
            Cylinder.next_inspection_date < date.today(),
            Cylinder.status == CylinderStatus.ACTIVE,
        ).count()
        repair_cost = db.query(Repair).filter(
            Repair.fleet_id == fleet.id
        ).with_entities(db.func.sum(Repair.cost)).scalar() or 0

        stock_items = db.query(StockItem).filter(StockItem.fleet_id == fleet.id).count()
        low_stock = db.query(StockItem).filter(
            StockItem.fleet_id == fleet.id,
            StockItem.quantity <= StockItem.min_quantity,
        ).count()

        results.append({
            "fleet_id": fleet.id,
            "fleet_name": fleet.name,
            "buses": buses,
            "cylinders": cylinders,
            "overdue_inspections": overdue,
            "total_repair_cost": float(repair_cost),
            "stock_items": stock_items,
            "low_stock_items": low_stock,
        })
    return results


@router.get("/reports/upcoming-inspections")
def upcoming_inspections(
    days: int = 90,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Cylinders with inspection due within `days` days."""
    today = date.today()
    deadline = today + date.resolution * days
    q = scope_query(db, Cylinder, user)
    cylinders = q.filter(
        Cylinder.next_inspection_date <= deadline,
        Cylinder.next_inspection_date >= today,
        Cylinder.status == CylinderStatus.ACTIVE,
    ).order_by(Cylinder.next_inspection_date).all()
    return cylinders


# ── Notification Settings ────────────────────────────────────────────────────

@router.get("/notifications/settings")
def get_notification_settings(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    ns = db.query(NotificationSettings).filter(
        NotificationSettings.user_id == user.id
    ).first()
    if not ns:
        ns = NotificationSettings(user_id=user.id)
        db.add(ns)
        db.commit()
    return ns


@router.put("/notifications/settings")
def update_notification_settings(
    notify_1month: bool = True,
    notify_1week: bool = True,
    notify_3months: bool = True,
    notify_overdue: bool = True,
    telegram_bot_token: str | None = None,
    telegram_chat_id: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    ns = db.query(NotificationSettings).filter(
        NotificationSettings.user_id == user.id
    ).first()
    if not ns:
        ns = NotificationSettings(user_id=user.id)
        db.add(ns)
    ns.notify_1month = notify_1month
    ns.notify_1week = notify_1week
    ns.notify_3months = notify_3months
    ns.notify_overdue = notify_overdue
    if telegram_bot_token is not None:
        ns.telegram_bot_token = telegram_bot_token
    if telegram_chat_id is not None:
        ns.telegram_chat_id = telegram_chat_id
    db.commit()
    return ns


# ── Excel Export ─────────────────────────────────────────────────────────────

@router.get("/reports/export")
def export_excel(
    fleet_id: int | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export cylinders data as Excel file."""
    import io, base64
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from datetime import date

    q = scope_query(db, Cylinder, user)
    if fleet_id:
        q = q.filter(Cylinder.fleet_id == fleet_id)
    cylinders = q.order_by(Cylinder.number).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Баллоны"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["Номер", "Клеймо", "Серийный №", "Тип газа", "Дата изгот.", "Объём, л",
               "Раб. давление", "Пробное давление", "Масса, кг", "Статус",
               "След. освид.", "Автопарк", "Автобус"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row, cyl in enumerate(cylinders, 2):
        fleet_name = cyl.fleet.name if cyl.fleet else ""
        bus_gosnomer = cyl.bus.gosnomer if cyl.bus else ""
        vals = [
            cyl.number, cyl.stamp or "", cyl.serial_number or "", cyl.gas_type or "",
            str(cyl.manufactured_date) if cyl.manufactured_date else "",
            cyl.capacity_liters or "", cyl.working_pressure or "", cyl.test_pressure or "",
            cyl.tare_weight or "", cyl.status or "",
            str(cyl.next_inspection_date) if cyl.next_inspection_date else "",
            fleet_name, bus_gosnomer
        ]
        overdue = cyl.next_inspection_date and cyl.next_inspection_date < date.today()
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = thin_border
            if overdue:
                cell.font = Font(color="DC3545")

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 16
    ws.column_dimensions['L'].width = 16
    ws.column_dimensions['M'].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    b64 = base64.b64encode(buf.read()).decode()

    return {"filename": f"cylinders_{date.today().isoformat()}.xlsx", "data": b64}


# ── Dashboard Stats ──────────────────────────────────────────────────────────

@router.get("/dashboard")
def dashboard(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    today = date.today()
    fleets_q = scope_query(db, Fleet, user)

    total_buses = scope_query(db, Bus, user).count()
    total_cylinders = scope_query(db, Cylinder, user).count()
    active_cylinders = scope_query(db, Cylinder, user).filter(
        Cylinder.status == CylinderStatus.ACTIVE
    ).count()
    overdue = scope_query(db, Cylinder, user).filter(
        Cylinder.next_inspection_date < today,
        Cylinder.status == CylinderStatus.ACTIVE,
    ).count()

    # Upcoming inspections in next 30 days
    d30 = today + date.resolution * 30
    upcoming = scope_query(db, Cylinder, user).filter(
        Cylinder.next_inspection_date <= d30,
        Cylinder.next_inspection_date >= today,
        Cylinder.status == CylinderStatus.ACTIVE,
    ).count()

    low_stock = scope_query(db, StockItem, user).filter(
        StockItem.quantity <= StockItem.min_quantity
    ).count()

    return {
        "total_fleets": fleets_q.count(),
        "total_buses": total_buses,
        "total_cylinders": total_cylinders,
        "active_cylinders": active_cylinders,
        "overdue_inspections": overdue,
        "upcoming_30days": upcoming,
        "low_stock_items": low_stock,
    }
